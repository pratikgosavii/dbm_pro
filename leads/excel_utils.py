import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from .models import Lead, LeadActivity

def import_leads_from_excel(excel_file, user):
    """
    Import leads from an Excel file
    
    Args:
        excel_file: Excel file uploaded by user
        user: The user initiating the import
        
    Returns:
        int: Number of leads imported
    """
    # Read Excel file into DataFrame
    try:
        df = pd.read_excel(excel_file)
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")
    
    # Validate required columns
    required_columns = ['first_name', 'last_name', 'email']
    missing_columns = [col for col in required_columns if col.lower() not in [c.lower() for c in df.columns]]
    
    if missing_columns:
        raise Exception(f"Missing required columns: {', '.join(missing_columns)}")
    
    # Standardize column names (case-insensitive)
    column_mapping = {}
    for col in df.columns:
        if col.lower() == 'first_name':
            column_mapping[col] = 'first_name'
        elif col.lower() == 'last_name':
            column_mapping[col] = 'last_name'
        elif col.lower() == 'email':
            column_mapping[col] = 'email'
        elif col.lower() in ['phone', 'phone_number']:
            column_mapping[col] = 'phone'
        elif col.lower() in ['company', 'company_name']:
            column_mapping[col] = 'company'
        elif col.lower() in ['job_title', 'title', 'position']:
            column_mapping[col] = 'job_title'
        elif col.lower() in ['notes', 'comments']:
            column_mapping[col] = 'notes'
    
    df = df.rename(columns=column_mapping)
    
    # Import leads
    imported_count = 0
    for _, row in df.iterrows():
        # Skip if email is missing
        if pd.isna(row.get('email', None)):
            continue
        
        # Check if lead already exists with this email
        existing_lead = Lead.objects.filter(email=row['email']).first()
        if existing_lead:
            # Update existing lead with new information
            existing_lead.first_name = row.get('first_name', existing_lead.first_name)
            existing_lead.last_name = row.get('last_name', existing_lead.last_name)
            existing_lead.phone = row.get('phone', existing_lead.phone)
            existing_lead.company = row.get('company', existing_lead.company)
            existing_lead.job_title = row.get('job_title', existing_lead.job_title)
            existing_lead.notes = row.get('notes', existing_lead.notes)
            existing_lead.save()
            
            # Create activity log
            LeadActivity.objects.create(
                lead=existing_lead,
                user=user,
                activity_type='Updated',
                description='Lead updated from Excel import'
            )
        else:
            # Create new lead
            new_lead = Lead(
                first_name=row.get('first_name', ''),
                last_name=row.get('last_name', ''),
                email=row['email'],
                phone=row.get('phone', ''),
                company=row.get('company', ''),
                job_title=row.get('job_title', ''),
                notes=row.get('notes', ''),
                status=Lead.NEW,
                source=Lead.EXCEL_IMPORT
            )
            new_lead.save()
            
            # Create activity log
            LeadActivity.objects.create(
                lead=new_lead,
                user=user,
                activity_type='Imported',
                description='Lead imported from Excel'
            )
            
            imported_count += 1
    
    return imported_count

def export_leads_to_excel(leads_queryset):
    """
    Export leads to Excel file
    
    Args:
        leads_queryset: QuerySet of leads to export
        
    Returns:
        bytes: Excel file content as bytes
    """
    # Create workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Leads"
    
    # Add headers
    headers = [
        'First Name', 'Last Name', 'Email', 'Phone', 'Company', 
        'Job Title', 'Status', 'Source', 'Notes', 'Assigned To',
        'Created At', 'Updated At'
    ]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    # Add data
    for row_num, lead in enumerate(leads_queryset, 2):
        worksheet.cell(row=row_num, column=1, value=lead.first_name)
        worksheet.cell(row=row_num, column=2, value=lead.last_name)
        worksheet.cell(row=row_num, column=3, value=lead.email)
        worksheet.cell(row=row_num, column=4, value=lead.phone)
        worksheet.cell(row=row_num, column=5, value=lead.company)
        worksheet.cell(row=row_num, column=6, value=lead.job_title)
        worksheet.cell(row=row_num, column=7, value=lead.get_status_display())
        worksheet.cell(row=row_num, column=8, value=lead.get_source_display())
        worksheet.cell(row=row_num, column=9, value=lead.notes)
        worksheet.cell(row=row_num, column=10, value=str(lead.assigned_to) if lead.assigned_to else '')
        worksheet.cell(row=row_num, column=11, value=lead.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        worksheet.cell(row=row_num, column=12, value=lead.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Save to in-memory file
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()
